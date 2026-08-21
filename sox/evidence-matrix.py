#!/usr/bin/env python3
"""
SOX ITGC Evidence Collection Matrix
=========================================
Maps every ITGC (see itgc_catalog.py) to the specific evidence artifacts a
tester or auditor needs to support design and operating-effectiveness
testing: what the evidence is, how/where to collect it, who owns producing
it, and the required retention period.

Outputs
-------
  outputs/evidence-matrix.xlsx
  outputs/evidence-matrix.md

Run:  python3 sox/evidence-matrix.py
"""

from __future__ import annotations

import os
import sys

from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(__file__))
from itgc_catalog import CATALOG_BY_ID  # noqa: E402
from xlsx_style import (  # noqa: E402
    add_table, autosize_columns, style_data_rows, style_header_row, style_title_block,
)

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")
os.makedirs(OUT_DIR, exist_ok=True)

RETENTION = "7 years (SOX Section 802 / SEC Rule 17a-4 alignment)"

# evidence_id, control_id, description, how_to_collect, owner, storage_location
EVIDENCE = [
    ("EV-001", "SOX-ITGC-001", "New-access request tickets with system/data-owner approval",
     "Export from ITSM ticketing system (e.g., ServiceNow), filtered by request type = "
     "New Access, for the testing period.",
     "IT Security / IAM Team", "ITSM archive / GRC evidence repository"),
    ("EV-002", "SOX-ITGC-001", "Provisioning confirmation showing access granted matches request",
     "Export from IAM platform or target-system provisioning/audit log.",
     "IT Security / IAM Team", "IAM platform export / GRC evidence repository"),

    ("EV-003", "SOX-ITGC-002", "HR termination report for the period",
     "Export from HRIS (e.g., Workday) termination module.",
     "HR / People Operations", "HRIS export / GRC evidence repository"),
    ("EV-004", "SOX-ITGC-002", "Access deactivation timestamps per terminated user",
     "Export from IAM deprovisioning workflow or target-system audit log.",
     "IT Security / IAM Team", "IAM platform export / GRC evidence repository"),

    ("EV-005", "SOX-ITGC-003", "Point-in-time privileged/admin account listing per system",
     "System-generated report from each in-scope system or PAM tool.",
     "IT Security", "PAM tool export / GRC evidence repository"),
    ("EV-006", "SOX-ITGC-003", "Privileged-access request & approval documentation",
     "Export from ITSM ticketing system, filtered by request type = Privileged Access.",
     "IT Security", "ITSM archive / GRC evidence repository"),
    ("EV-007", "SOX-ITGC-003", "MFA enforcement and privileged-session logging configuration",
     "Configuration export/screenshot from IAM and SIEM platforms.",
     "IT Security", "GRC evidence repository"),

    ("EV-008", "SOX-ITGC-004", "Approved SoD conflict rule set / matrix",
     "Signed-off document maintained jointly by IT Security and the Controller's Office.",
     "IT Security / Controller's Office", "GRC / policy repository"),
    ("EV-009", "SOX-ITGC-004", "SoD conflict analysis report for the period",
     "Automated output from the GRC/SoD analysis tool run against active user population.",
     "IT Security", "GRC evidence repository"),
    ("EV-010", "SOX-ITGC-004", "Mitigating/compensating control evidence for flagged conflicts",
     "Management review reports and sign-offs for each flagged conflict.",
     "Controller's Office", "GRC evidence repository"),

    ("EV-011", "SOX-ITGC-005", "UAR access listing with reviewer sign-off",
     "Export from UAR tool, or system access listing plus documented email/e-signature "
     "approval.",
     "Business System Owner", "GRC evidence repository"),
    ("EV-012", "SOX-ITGC-005", "Remediation evidence for access flagged for removal",
     "IAM deprovisioning log filtered to users flagged during the UAR.",
     "IT Security / IAM Team", "IAM platform export / GRC evidence repository"),

    ("EV-013", "SOX-ITGC-006", "Change tickets with justification, test evidence, and approval",
     "Export from the change management system (e.g., Jira, ServiceNow) for the period.",
     "IT Change Management", "Change management system archive"),
    ("EV-014", "SOX-ITGC-006", "Deployment logs identifying the individual who deployed",
     "Export from the CI/CD or deployment tool.",
     "Application Development / DevOps", "CI/CD system export"),

    ("EV-015", "SOX-ITGC-007", "Emergency change tickets with retroactive approval timestamp",
     "Export from the change management system, filtered by change type = Emergency.",
     "IT Change Management", "Change management system archive"),
    ("EV-016", "SOX-ITGC-007", "Reconciliation of emergency changes to the standard change register",
     "Reconciliation report/spreadsheet prepared by IT Change Management.",
     "IT Change Management", "GRC evidence repository"),

    ("EV-017", "SOX-ITGC-008", "Approved security configuration baseline and current-state export",
     "Approved baseline document plus configuration export from each in-scope system.",
     "IT Security", "GRC / policy repository"),
    ("EV-018", "SOX-ITGC-008", "SIEM log ingestion and alert review/triage evidence",
     "SIEM dashboard export and tickets for reviewed/triaged alerts.",
     "IT Security / SOC", "SIEM export / GRC evidence repository"),

    ("EV-019", "SOX-ITGC-009", "Batch job completion and failure-investigation logs",
     "Export from the job scheduler (e.g., Control-M, Autosys) completion report.",
     "IT Operations", "Job scheduler export / GRC evidence repository"),
    ("EV-020", "SOX-ITGC-009", "Backup completion logs and quarterly restore-test results",
     "Backup platform completion report plus restore-test evidence/screenshots.",
     "IT Operations", "Backup platform export / GRC evidence repository"),
]

HEADERS = ["Evidence ID", "Control ID", "Control Title", "Evidence Description",
           "How to Collect", "Evidence Owner", "Retention Period", "Storage Location"]


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence Matrix"
    ws.sheet_view.showGridLines = False

    header_row = style_title_block(
        ws,
        "SOX ITGC — Evidence Collection Matrix",
        "SOX Compliance Package | Confidential - Internal Use | Generated by evidence-matrix.py",
        n_cols=len(HEADERS),
    )
    for col, h in enumerate(HEADERS, start=1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, len(HEADERS))

    r = header_row + 1
    first_data_row = r
    for eid, cid, desc, how, owner, storage in EVIDENCE:
        ctl_title = CATALOG_BY_ID[cid].title
        ws.cell(row=r, column=1, value=eid)
        ws.cell(row=r, column=2, value=cid)
        ws.cell(row=r, column=3, value=ctl_title)
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value=how)
        ws.cell(row=r, column=6, value=owner)
        ws.cell(row=r, column=7, value=RETENTION)
        ws.cell(row=r, column=8, value=storage)
        r += 1
    last_data_row = r - 1

    style_data_rows(ws, first_data_row, last_data_row, len(HEADERS))
    autosize_columns(ws, {1: 12, 2: 14, 3: 40, 4: 44, 5: 44, 6: 28, 7: 34, 8: 36})
    add_table(ws, "EvidenceMatrix", f"A{header_row}:H{last_data_row}")
    return wb


def write_markdown(path: str):
    lines = [
        "# SOX ITGC — Evidence Collection Matrix",
        "",
        "*SOX Compliance Package | Confidential - Internal Use*",
        "",
        f"{len(EVIDENCE)} evidence artifacts mapped across "
        f"{len({e[1] for e in EVIDENCE})} IT General Controls. Standard retention period: "
        f"**{RETENTION}**.",
        "",
        "| Evidence ID | Control | Description | How to Collect | Owner | Storage |",
        "|---|---|---|---|---|---|",
    ]
    for eid, cid, desc, how, owner, storage in EVIDENCE:
        ctl_title = CATALOG_BY_ID[cid].title
        lines.append(f"| {eid} | {cid}: {ctl_title} | {desc} | {how} | {owner} | {storage} |")
    with open(path, "w") as f:
        f.write("\n".join(lines) + "\n")


def main():
    wb = build_workbook()
    xlsx_path = os.path.join(OUT_DIR, "evidence-matrix.xlsx")
    wb.save(xlsx_path)
    md_path = os.path.join(OUT_DIR, "evidence-matrix.md")
    write_markdown(md_path)
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {md_path}")
    print(f"{len(EVIDENCE)} evidence items across {len({e[1] for e in EVIDENCE})} controls")


if __name__ == "__main__":
    main()
