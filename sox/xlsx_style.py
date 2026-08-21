"""
Shared openpyxl styling helpers for the SOX Compliance Package workbooks.
Keeps every generator script (COSO mapping, IT General Controls, Evidence
Matrix, Control Design Template) visually consistent: same header banner,
same risk/status color scale, same column-sizing/freeze-pane behavior.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------- palette
NAVY = "1F3864"
NAVY_LIGHT = "2E4E8C"
WHITE = "FFFFFF"
GREY_LIGHT = "F2F2F2"
BORDER_GREY = "BFBFBF"

RISK_FILL = {
    "Low": "C6E0B4",       # green
    "Medium": "FFE699",    # amber
    "High": "F4B183",      # orange
    "Critical": "E06666",  # red
}
RISK_FONT_COLOR = {
    "Low": "375623",
    "Medium": "7F6000",
    "High": "833C00",
    "Critical": "FFFFFF",
}

STATUS_FILL = {
    "Not Started": "D9D9D9",
    "Planned": "FFE699",
    "In Progress": "9DC3E6",
    "Implemented": "C6E0B4",
    "Not Effective": "E06666",
    "Not Applicable": "D9D9D9",
}

thin = Side(style="thin", color=BORDER_GREY)
CELL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
BODY_FONT = Font(name="Calibri", size=10)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_title_block(ws, title: str, subtitle: str, n_cols: int, row: int = 1):
    """Write a merged title + subtitle banner across the top n_cols columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=n_cols)
    sub = ws.cell(row=row + 1, column=1, value=subtitle)
    sub.font = SUBTITLE_FONT
    ws.row_dimensions[row + 1].height = 16
    return row + 3  # next free row for the header


def style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = CELL_BORDER
    ws.row_dimensions[row].height = 32
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def style_data_rows(ws, first_row: int, last_row: int, n_cols: int, risk_col: int | None = None,
                     status_col: int | None = None):
    for r in range(first_row, last_row + 1):
        stripe = GREY_LIGHT if (r - first_row) % 2 == 1 else WHITE
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.font = BODY_FONT
            c.alignment = WRAP_TOP
            c.border = CELL_BORDER
            if not c.fill or c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = PatternFill("solid", fgColor=stripe)
        if risk_col:
            rc = ws.cell(row=r, column=risk_col)
            level = str(rc.value)
            if level in RISK_FILL:
                rc.fill = PatternFill("solid", fgColor=RISK_FILL[level])
                rc.font = Font(name="Calibri", size=10, bold=True, color=RISK_FONT_COLOR[level])
                rc.alignment = CENTER
        if status_col:
            sc = ws.cell(row=r, column=status_col)
            status = str(sc.value)
            if status in STATUS_FILL:
                sc.fill = PatternFill("solid", fgColor=STATUS_FILL[status])
                sc.alignment = CENTER


def autosize_columns(ws, widths: dict[int, int]):
    """widths: {col_index (1-based): width}"""
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def add_table(ws, name: str, ref: str):
    """Register an Excel structured Table (adds filter dropdowns) without
    re-styling — we already hand-style headers/rows for the banded look."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(tbl)


def write_legend_sheet(wb, risk_definitions: dict[str, str]):
    ws = wb.create_sheet("Legend")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Risk Rating Legend"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    row = 3
    ws.cell(row=row, column=1, value="Rating").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Definition").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 1
    for level, definition in risk_definitions.items():
        c1 = ws.cell(row=row, column=1, value=level)
        c1.fill = PatternFill("solid", fgColor=RISK_FILL[level])
        c1.font = Font(bold=True, color=RISK_FONT_COLOR[level])
        c1.alignment = CENTER
        c1.border = CELL_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c2 = ws.cell(row=row, column=2, value=definition)
        c2.font = BODY_FONT
        c2.alignment = WRAP_TOP
        c2.border = CELL_BORDER
        row += 1
    autosize_columns(ws, {1: 14, 2: 60, 3: 20})
    return ws
